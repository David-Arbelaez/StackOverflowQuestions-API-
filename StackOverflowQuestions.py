import os
import sys
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

if getattr(sys, 'frozen', False):  # Si está congelado por PyInstaller
    base_path = os.path.dirname(sys.executable)
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

response = requests.get('https://api.stackexchange.com/2.3/questions?order=desc&sort=activity&site=stackoverflow')
data = response.json()['items']
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Unanswered questions"

ws['B2'] = "Title"
ws['C3'] = "Link"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
alignment = Alignment(horizontal="center", wrap_text=True)

ws['B2'].font = header_font
ws['B2'].fill = header_fill
ws['B2'].alignment = alignment

ws['C3'].font = header_font
ws['C3'].fill = header_fill
ws['C3'].alignment = alignment

row = 3
for item in data:
    if item['answer_count'] == 0:
        title_cell = ws.cell(row=row, column=2)
        link_cell = ws.cell(row=row, column=3)

        title_cell.value = item['title']
        title_cell.alignment = Alignment(wrap_text=True)

        link_cell.value = "link"
        link_cell.hyperlink = item['link']
        link_cell.style = 'Hyperlink'
        link_cell.alignment = Alignment(horizontal="center")

        row += 1

ws.column_dimensions['B'].width = 100
ws.column_dimensions['C'].width = 15

output_folder = os.path.join(base_path, "output")
os.makedirs(output_folder, exist_ok=True)

output_path = os.path.join(output_folder, "Unanswered_questions.xlsx")
wb.save(output_path)

os.startfile(output_path)
